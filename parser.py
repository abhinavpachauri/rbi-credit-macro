"""
RBI SIBC (Sector/Industry-wise Bank Credit) Excel Parser
=========================================================
Parses the monthly SIBC report published by RBI.
Filename convention: SIBC{DDMMYYYY}.xlsx

Outputs two artefacts:
  1. <stem>_wide.csv  — one row per sector/industry, all columns preserved
  2. <stem>_long.csv  — tidy long format: (date, code, sector, outstanding_cr)
                        suitable for time-series consolidation
"""

import re
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd


# ---------------------------------------------------------------------------
# Date parsing helpers
# ---------------------------------------------------------------------------

_MONTH_MAP = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}

def parse_rbi_date(text: str) -> datetime | None:
    """Parse dates like '26.Jan,2024' or '31.Jan,2026'."""
    if not text:
        return None
    m = re.match(r"(\d{1,2})\.(\w{3}),(\d{4})", str(text).strip())
    if not m:
        return None
    day, mon, year = int(m.group(1)), _MONTH_MAP.get(m.group(2).lower()), int(m.group(3))
    if mon is None:
        return None
    return datetime(year, mon, day)

def parse_filename_date(path: Path) -> datetime | None:
    """
    Extract the publication date from an SIBC filename.

    Handles two conventions used by RBI:
      SIBC{DDMMYYYY}.xlsx      e.g. SIBC27022026   → 2026-02-27
      PR####SIBC{DDMMYY}.xlsx  e.g. PR2019SIBC300126 → 2026-01-30
                                    (DDMMYY at end, YY interpreted as 20YY)
    """
    stem = path.stem

    # Priority 1: 8-digit DDMMYYYY (e.g. 27022026)
    m = re.search(r"(\d{2})(\d{2})(\d{4})", stem)
    if m:
        day, mon, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return datetime(year, mon, day)
        except ValueError:
            pass  # fall through to next pattern

    # Priority 2: 6-digit DDMMYY at the end of the stem (e.g. ...300126)
    m = re.search(r"(\d{2})(\d{2})(\d{2})$", stem)
    if m:
        day, mon, yy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        year = 2000 + yy   # safe until 2099
        try:
            return datetime(year, mon, day)
        except ValueError:
            pass

    return None


# ---------------------------------------------------------------------------
# Hierarchy / code detection helpers
# ---------------------------------------------------------------------------

# Patterns for code detection (applied to stripped sector name)
_CODE_PATTERNS = [
    # Roman numerals: I., II., III.
    (re.compile(r"^(I{1,3}V?|VI{0,3}|IX|IV)\.\s+(.+)$"), 0),
    # x.y.z pattern (e.g. 3.7.1, 2.18.1)
    (re.compile(r"^(\d+\.\d+\.\d+)\.\s+(.+)$"), 3),
    # x.y pattern (e.g. 2.1, 3.10)
    (re.compile(r"^(\d+\.\d+)\.\s+(.+)$"), 2),
    # Digit only (e.g. 1., 2., 3., 4.)
    (re.compile(r"^(\d+)\.\s+(.+)$"), 1),
    # Priority sector sub-items: (i), (ii), ...
    (re.compile(r"^\(([ivxlcdm]+)\)\s+(.+)$", re.IGNORECASE), 2),
]

_FOOTNOTE_RE   = re.compile(r"\d+\s*$")          # trailing footnote numbers
_INLINE_FN_RE  = re.compile(r"\)\d+")            # footnote digit after ")" e.g. "NBFCs)2"
_OF_WHICH_RE   = re.compile(r"\s+of\s+which,?\s*$", re.IGNORECASE)
_MULTISPACE_RE = re.compile(r"  +")              # multiple spaces inside name


def clean_sector_name(name: str) -> str:
    """Strip footnote artefacts and normalise whitespace."""
    name = _INLINE_FN_RE.sub(")", name)           # "NBFCs)2 of which," → "NBFCs) of which,"
    name = _OF_WHICH_RE.sub("", name).strip()     # strip " of which," suffix
    name = _FOOTNOTE_RE.sub("", name).strip()     # strip trailing standalone digits
    name = name.rstrip(",").strip()               # strip trailing comma
    name = _MULTISPACE_RE.sub(" ", name)
    return name


def detect_code_and_level(raw_name: str):
    """
    Returns (code, cleaned_name, level) for a sector/industry cell value.
    raw_name may have leading whitespace (indentation in the original sheet).
    """
    # Preserve leading spaces to infer relative indentation
    leading_spaces = len(raw_name) - len(raw_name.lstrip(" "))
    name = raw_name.strip()

    for pattern, level in _CODE_PATTERNS:
        m = pattern.match(name)
        if m:
            code = m.group(1)
            cleaned = clean_sector_name(m.group(2).strip())
            # Sub-items with leading spaces bump their level up
            if leading_spaces >= 4:
                level = max(level, 3)
            elif leading_spaces >= 2:
                level = max(level, 2)
            return code, cleaned, level

    # No code found → either a header row, a notes row, or a label
    return None, clean_sector_name(name), -1


# ---------------------------------------------------------------------------
# Notes / separator row detection
# ---------------------------------------------------------------------------

_NOTES_PREFIXES = (
    "note", "notes:", "(1)", "(2)", "(3)", "1 ", "2 ", "3 ", "4 ", "5 ",
    "6 ", "7 ",
)

def is_notes_row(name: str) -> bool:
    low = name.strip().lower()
    return any(low.startswith(p) for p in _NOTES_PREFIXES)

def is_separator_row(row_values) -> bool:
    """A row where all numeric columns are None/empty."""
    return all(v is None for v in row_values[1:])


# ---------------------------------------------------------------------------
# Parent-code derivation
# ---------------------------------------------------------------------------

# The only cross-level hardcoded link: level-1 numeric sectors roll up into III.
# (I = II + III; II = Food Credit standalone; III = Non-food Credit = 1+2+3+4)
_LEVEL1_PARENT_CODE      = "III"
_LEVEL1_PARENT_STATEMENT = "Statement 1"

# Statement 2 is an industry drill-down of sector 2 (Industry) in Statement 1.
_STMT2_ROOT_CODE      = "2"
_STMT2_ROOT_STATEMENT = "Statement 1"


def derive_parent(code, level: int, statement: str, is_priority_sector_memo: bool):
    """
    Return (parent_code, parent_statement) for a single row.

    Rules
    -----
    Priority-sector memo items  → (None, None)  — memo block, outside main tree
    Level -1  (total rows)      → (None, None)
    Level  0  (I, II, III)      → (None, None)  — roots

    Statement 1, level 1 (1–4) → ("III", "Statement 1")   [hardcoded domain link]
    Statement 1, level 2 (x.y) → (x,    "Statement 1")    [drop last segment]
    Statement 1, level 3 (x.y.z)→(x.y,  "Statement 1")   [drop last segment]

    Statement 2, level 2 (2.x) → ("2",  "Statement 1")    [cross-statement root]
    Statement 2, level 3 (2.x.y)→(2.x,  "Statement 2")   [drop last segment]
    """
    if is_priority_sector_memo:
        return None, None
    if level <= 0 or code is None:
        return None, None

    code_str = str(code)

    # ---- Statement 2 ----
    if statement == "Statement 2":
        if level == 2:
            # 2.x → parent is code "2" (Industry) in Statement 1
            return _STMT2_ROOT_CODE, _STMT2_ROOT_STATEMENT
        if level == 3 and "." in code_str:
            # 2.x.y → parent is 2.x in Statement 2
            parent = code_str.rsplit(".", 1)[0]
            return parent, "Statement 2"
        return None, None

    # ---- Statement 1 ----
    if level == 1:
        # Numeric sector codes 1, 2, 3, 4 → parent is III (Non-food Credit)
        return _LEVEL1_PARENT_CODE, _LEVEL1_PARENT_STATEMENT

    if "." in code_str:
        # x.y   → x      (level 2 → level 1)
        # x.y.z → x.y    (level 3 → level 2)
        parent = code_str.rsplit(".", 1)[0]
        return parent, "Statement 1"

    return None, None


def add_parent_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Vectorised wrapper: applies derive_parent row-wise and appends two columns."""
    parents = df.apply(
        lambda r: derive_parent(
            r["code"], r["level"], r["statement"], r["is_priority_sector_memo"]
        ),
        axis=1,
        result_type="expand",
    )
    parents.columns = ["parent_code", "parent_statement"]
    return pd.concat([df, parents], axis=1)


# Columns that must always be read as strings (codes can look numeric: "2", "3.10")
STRING_COLS = ["code", "parent_code", "statement", "parent_statement", "sector"]


def enforce_string_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Ensure code/sector columns stay as proper strings even after CSV round-trips.
    pandas infers float for columns like ["2", NaN, "3.10"] → "2.0", "3.1".
    We convert to str and replace the literal string "nan" back to None.
    """
    for col in STRING_COLS:
        if col in df.columns:
            df[col] = df[col].where(df[col].isna(), df[col].astype(str))
    return df


# ---------------------------------------------------------------------------
# Sheet parser
# ---------------------------------------------------------------------------

def parse_sheet(ws_data: list[tuple], sheet_name: str, report_date: datetime) -> pd.DataFrame:
    """
    ws_data : list of row tuples from openpyxl (values_only=True)
    Returns a tidy wide DataFrame.
    """
    # ---- locate header rows ----
    # Row 0: title, Row 1: unit, Row 2: col header row 1, Row 3: col header row 2
    # We need rows 3 (dates) and 4 ('%' indicators) to build column names.
    # Rows are 0-indexed here.

    header_row_dates = ws_data[3]   # e.g. ('', '26.Jan,2024', ...)
    # header_row_pct  = ws_data[4]  # only used to confirm % columns

    # Parse dates from the header
    date_labels = []
    for cell in header_row_dates[1:]:
        if cell and str(cell).strip():
            date_labels.append(str(cell).strip())

    # The sheet has: 5 outstanding cols + 2 YoY cols + 2 FY cols = 9 numeric cols
    # Map column indices (1-based from raw row) to semantic names
    n_outstanding = 5   # columns B-F
    n_yoy = 2           # columns G-H
    n_fy = 2            # columns I-J

    out_dates = date_labels[:n_outstanding]
    yoy_pairs = date_labels[n_outstanding: n_outstanding + n_yoy]
    fy_pairs  = date_labels[n_outstanding + n_yoy: n_outstanding + n_yoy + n_fy]

    # ---- build column schema ----
    # We'll name outstanding cols by their date, e.g. "2024-01-26"
    def fmt(label):
        d = parse_rbi_date(label)
        return d.strftime("%Y-%m-%d") if d else label.replace(" ", "_")

    out_col_names = [fmt(d) for d in out_dates]
    yoy_col_names = [f"yoy_pct__{p.replace(' ', '').replace('/', '__')}" for p in yoy_pairs]
    fy_col_names  = [f"fy_pct__{p.replace(' ', '').replace('/', '__')}" for p in fy_pairs]

    numeric_cols = out_col_names + yoy_col_names + fy_col_names

    # ---- parse data rows ----
    records = []
    in_priority_sector = False  # tracks when we enter the Priority Sector memo block
    for raw_row in ws_data[5:]:  # skip title+unit+2 header rows+% row
        first_cell = raw_row[0]
        if first_cell is None:
            continue

        name_str = str(first_cell)

        # Stop at notes
        if is_notes_row(name_str):
            break

        # Detect Priority Sector memo label row BEFORE separator check
        if "priority sector" in name_str.lower() and "(memo)" in name_str.lower():
            in_priority_sector = True
            continue  # skip the label row itself

        # Skip separator rows
        if is_separator_row(raw_row):
            continue

        code, clean_name, level = detect_code_and_level(name_str)

        # Skip un-parseable label rows
        if code is None and level == -1:
            has_data = any(isinstance(v, (int, float)) for v in raw_row[1:])
            if not has_data:
                continue

        numeric_values = list(raw_row[1: 1 + len(numeric_cols)])
        # Pad if fewer columns than expected
        while len(numeric_values) < len(numeric_cols):
            numeric_values.append(None)

        record = {
            "report_date": report_date.strftime("%Y-%m-%d") if report_date else None,
            "statement": sheet_name,
            "code": code,
            "sector": clean_name,
            "level": level,
            "is_priority_sector_memo": in_priority_sector,
        }
        for col, val in zip(numeric_cols, numeric_values):
            record[col] = val

        records.append(record)

    df = pd.DataFrame(records)
    df = add_parent_columns(df)
    df = enforce_string_cols(df)
    return df


# ---------------------------------------------------------------------------
# Long-format converter
# ---------------------------------------------------------------------------

def to_long(df_wide: pd.DataFrame) -> pd.DataFrame:
    """
    Melt outstanding columns into long format:
      report_date | statement | code | sector | level
      | is_priority_sector_memo | parent_code | parent_statement
      | date | outstanding_cr
    """
    # Identify outstanding columns (named as YYYY-MM-DD)
    date_cols = [c for c in df_wide.columns if re.match(r"\d{4}-\d{2}-\d{2}", c)]
    id_vars = [
        "report_date", "statement", "code", "sector", "level",
        "is_priority_sector_memo", "parent_code", "parent_statement",
    ]
    # Only keep id_vars that actually exist (guard against missing cols)
    id_vars = [c for c in id_vars if c in df_wide.columns]

    df_long = df_wide[id_vars + date_cols].melt(
        id_vars=id_vars,
        value_vars=date_cols,
        var_name="date",
        value_name="outstanding_cr",
    )
    df_long = df_long.dropna(subset=["outstanding_cr"])
    df_long = df_long.sort_values(["statement", "code", "date"]).reset_index(drop=True)
    return df_long


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def parse_file(filepath: str | Path) -> dict[str, pd.DataFrame]:
    """
    Parse a single SIBC Excel file.
    Returns {
        "stmt1_wide": DataFrame,
        "stmt2_wide": DataFrame,
        "stmt1_long": DataFrame,
        "stmt2_long": DataFrame,
    }
    """
    import openpyxl

    path = Path(filepath)
    report_date = parse_filename_date(path)
    if report_date is None:
        print(f"[WARN] Could not parse report date from filename: {path.name}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    results = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws_data = [tuple(row) for row in ws.iter_rows(values_only=True)]
        df_wide = parse_sheet(ws_data, sheet_name, report_date)
        df_long = to_long(df_wide)
        key = sheet_name.lower().replace(" ", "")
        results[f"{key}_wide"] = df_wide
        results[f"{key}_long"] = df_long

    wb.close()
    return results


def save_outputs(filepath: str | Path, out_dir: str | Path | None = None):
    """Parse and save CSV outputs next to the source file (or to out_dir)."""
    path = Path(filepath)
    out_dir = Path(out_dir) if out_dir else path.parent
    out_dir.mkdir(parents=True, exist_ok=True)

    results = parse_file(path)
    stem = path.stem

    for key, df in results.items():
        out_path = out_dir / f"{stem}_{key}.csv"
        df.to_csv(out_path, index=False)
        print(f"  Saved {out_path.name}  ({len(df)} rows)")

    return results


def read_parsed_csv(csv_path: str | Path) -> pd.DataFrame:
    """
    Read a CSV produced by this parser with correct dtypes.
    Always use this instead of pd.read_csv() directly — otherwise pandas
    misinterprets numeric-looking codes ('2', '3.10') as floats.
    """
    path = Path(csv_path)
    # Read all STRING_COLS as str to prevent float inference
    dtype_map = {col: str for col in STRING_COLS}
    df = pd.read_csv(path, dtype=dtype_map)
    # Replace literal "nan" strings (from None → CSV → str read-back) with actual NaN
    for col in STRING_COLS:
        if col in df.columns:
            df[col] = df[col].replace("nan", pd.NA)
    return df


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python parser.py <SIBC_FILE.xlsx> [output_dir]")
        sys.exit(1)

    xlsx_path = sys.argv[1]
    out_dir   = sys.argv[2] if len(sys.argv) > 2 else None

    print(f"\nParsing: {xlsx_path}")
    results = save_outputs(xlsx_path, out_dir)

    # Quick summary
    print("\n--- Summary ---")
    for key, df in results.items():
        print(f"  {key}: {df.shape[0]} rows × {df.shape[1]} cols")
        if not df.empty:
            print(f"    Columns: {list(df.columns)}")
